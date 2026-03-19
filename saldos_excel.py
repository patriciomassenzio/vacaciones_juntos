import re
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
EXCEL_DIR = BASE_DIR / "excels"

ARCHIVOS_SALDOS = [
    EXCEL_DIR / "Anexo 2_Amazonas Bagua ROL VACAC_2025_v3.xlsx",
    EXCEL_DIR / "Anexo 2_Amazonas Bagua ROL VACAC_2026.xlsx",
]

DIAS_MAXIMOS_POR_PERIODO = 30


def normalizar_dni(dni):
    if dni is None:
        return ""
    return re.sub(r"\D", "", str(dni).strip())


def extraer_periodo_mencionado(texto):
    if not texto:
        return None

    match = re.search(r"(20\d{2}\s*-\s*20\d{2})", texto)
    if match:
        return match.group(1).replace(" ", "")
    return None


def buscar_fila_encabezado(ws):
    for fila in range(1, min(ws.max_row, 20) + 1):
        valores = [ws.cell(fila, col).value for col in range(1, ws.max_column + 1)]
        if any(v and "DNI" in str(v).upper() for v in valores):
            return fila
    return None


def convertir_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def dias_inclusivos(desde, hasta):
    if not desde or not hasta:
        return 0
    return (hasta - desde).days + 1


def leer_registros_archivo(path_archivo):
    if not path_archivo.exists():
        raise FileNotFoundError(f"No encontré el archivo Excel: {path_archivo}")

    wb = load_workbook(path_archivo, data_only=True)

    if "UTAB" not in wb.sheetnames:
        raise ValueError(f"El archivo no tiene la hoja UTAB: {path_archivo.name}")

    ws = wb["UTAB"]
    fila_header = buscar_fila_encabezado(ws)

    if not fila_header:
        raise ValueError(f"No encontré encabezados en la hoja UTAB de {path_archivo.name}")

    registros = []

    for fila in range(fila_header + 1, ws.max_row + 1):
        dni = ws.cell(fila, 3).value
        nombre = ws.cell(fila, 4).value
        periodo = ws.cell(fila, 7).value
        desde = ws.cell(fila, 8).value
        hasta = ws.cell(fila, 9).value
        observacion = ws.cell(fila, 12).value

        dni_norm = normalizar_dni(dni)

        if not dni_norm:
            continue

        desde_fecha = convertir_fecha(desde)
        hasta_fecha = convertir_fecha(hasta)

        total_dias = dias_inclusivos(desde_fecha, hasta_fecha)

        registros.append({
            "archivo": path_archivo.name,
            "dni": dni_norm,
            "nombre": str(nombre).strip() if nombre else "",
            "periodo": str(periodo).strip() if periodo else "",
            "desde": desde_fecha,
            "hasta": hasta_fecha,
            "dias": total_dias,
            "observacion": str(observacion).strip() if observacion else ""
        })

    return registros


def cargar_todos_los_registros():
    todos = []
    for archivo in ARCHIVOS_SALDOS:
        todos.extend(leer_registros_archivo(archivo))
    return todos


def obtener_registros_por_dni(dni):
    dni_norm = normalizar_dni(dni)
    todos = cargar_todos_los_registros()
    return [r for r in todos if r["dni"] == dni_norm]


def obtener_resumen_saldos(dni):
    registros = obtener_registros_por_dni(dni)

    resumen = {}

    for r in registros:
        periodo = r["periodo"]

        if periodo not in resumen:
            resumen[periodo] = {
                "dni": r["dni"],
                "nombre": r["nombre"],
                "periodo": periodo,
                "dias_maximos": DIAS_MAXIMOS_POR_PERIODO,
                "dias_usados": 0,
                "saldo_disponible": 0,
                "detalle": []
            }

        resumen[periodo]["dias_usados"] += r["dias"]
        resumen[periodo]["detalle"].append({
            "desde": str(r["desde"]) if r["desde"] else None,
            "hasta": str(r["hasta"]) if r["hasta"] else None,
            "dias": r["dias"],
            "observacion": r["observacion"],
            "archivo": r["archivo"]
        })

    for periodo in resumen:
        usados = resumen[periodo]["dias_usados"]
        resumen[periodo]["saldo_disponible"] = DIAS_MAXIMOS_POR_PERIODO - usados

    return sorted(resumen.values(), key=lambda x: x["periodo"])


def validar_saldo_vacacional(dni, dias_solicitados, periodo_mencionado=None):
    resumenes = obtener_resumen_saldos(dni)

    if not resumenes:
        return False, "No encontré saldo vacacional para ese DNI en los archivos Excel.", None

    if periodo_mencionado:
        encontrados = [r for r in resumenes if r["periodo"] == periodo_mencionado]

        if not encontrados:
            periodos = ", ".join(r["periodo"] for r in resumenes)
            return False, f"No encontré el período {periodo_mencionado} para ese DNI. Períodos disponibles: {periodos}.", None

        elegido = encontrados[0]

        if elegido["saldo_disponible"] < dias_solicitados:
            return False, (
                f"No tenés saldo suficiente en el período {elegido['periodo']}. "
                f"Saldo disponible: {elegido['saldo_disponible']} días. "
                f"Días solicitados: {dias_solicitados}."
            ), None

        return True, (
            f"Saldo validado en el período {elegido['periodo']}. "
            f"Saldo disponible: {elegido['saldo_disponible']} días."
        ), elegido["periodo"]

    con_saldo = [r for r in resumenes if r["saldo_disponible"] > 0]

    if len(con_saldo) == 0:
        detalle = " | ".join(
            f"{r['periodo']}: saldo {r['saldo_disponible']} días"
            for r in resumenes
        )
        return False, f"No tenés saldo disponible. Detalle: {detalle}.", None

    if len(con_saldo) > 1:
        detalle = " | ".join(
            f"{r['periodo']}: saldo {r['saldo_disponible']} días"
            for r in con_saldo
        )
        return False, (
            "Tenés saldo en más de un período vacacional. "
            f"Indicá el período en tu mensaje, por ejemplo 2025-2026. Detalle: {detalle}."
        ), None

    elegido = con_saldo[0]

    if elegido["saldo_disponible"] < dias_solicitados:
        return False, (
            f"No tenés saldo suficiente en el período {elegido['periodo']}. "
            f"Saldo disponible: {elegido['saldo_disponible']} días. "
            f"Días solicitados: {dias_solicitados}."
        ), None

    return True, (
        f"Saldo validado en el período {elegido['periodo']}. "
        f"Saldo disponible: {elegido['saldo_disponible']} días."
    ), elegido["periodo"]