import re
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook

from crud_saldos import reemplazar_todos_los_saldos

BASE_DIR = Path(__file__).resolve().parent
EXCEL_DIR = BASE_DIR / "excels"

ARCHIVOS_SALDOS = [
    EXCEL_DIR / "Anexo 2_Amazonas Bagua ROL VACAC_2025_v3.xlsx",
    EXCEL_DIR / "Anexo 2_Amazonas Bagua ROL VACAC_2026.xlsx",
]

DIAS_MAXIMOS_POR_PERIODO = 30

def normalizar_dni(dni):
    if dni is None:
        return ""
    return re.sub(r"\D", "", str(dni).strip())

def buscar_fila_encabezado(ws):
    for fila in range(1, min(ws.max_row, 20) + 1):
        valores = [ws.cell(fila, col).value for col in range(1, ws.max_column + 1)]
        if any(v and "DNI" in str(v).upper() for v in valores):
            return fila
    return None

def convertir_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None

def dias_inclusivos(desde, hasta):
    if not desde or not hasta:
        return 0
    return (hasta - desde).days + 1

def leer_registros_archivo(path_archivo):
    if not path_archivo.exists():
        raise FileNotFoundError(f"No encontré el archivo Excel: {path_archivo}")

    wb = load_workbook(path_archivo, data_only=True)

    if "UTAB" not in wb.sheetnames:
        raise ValueError(f"El archivo no tiene la hoja UTAB: {path_archivo.name}")

    ws = wb["UTAB"]
    fila_header = buscar_fila_encabezado(ws)

    if not fila_header:
        raise ValueError(f"No encontré encabezados en la hoja UTAB de {path_archivo.name}")

    registros = []

    for fila in range(fila_header + 1, ws.max_row + 1):
        dni = ws.cell(fila, 3).value
        nombre = ws.cell(fila, 4).value
        periodo = ws.cell(fila, 7).value
        desde = ws.cell(fila, 8).value
        hasta = ws.cell(fila, 9).value

        dni_norm = normalizar_dni(dni)
        if not dni_norm:
            continue

        desde_fecha = convertir_fecha(desde)
        hasta_fecha = convertir_fecha(hasta)
        total_dias = dias_inclusivos(desde_fecha, hasta_fecha)

        registros.append({
            "dni": dni_norm,
            "nombres": str(nombre).strip() if nombre else "",
            "periodo_vacacional": str(periodo).strip() if periodo else "",
            "dias_maximos": DIAS_MAXIMOS_POR_PERIODO,
            "dias_usados": total_dias,
            "saldo_disponible": 0,
            "fuente_archivo": path_archivo.name
        })

    return registros

def agrupar_saldos():
    agrupado = {}

    for archivo in ARCHIVOS_SALDOS:
        if not archivo.exists():
            raise FileNotFoundError(f"No encontré el archivo: {archivo}")

        registros = leer_registros_archivo(archivo)

        for r in registros:
            clave = (r["dni"], r["periodo_vacacional"])

            if clave not in agrupado:
                agrupado[clave] = {
                    "dni": r["dni"],
                    "nombres": r["nombres"],
                    "periodo_vacacional": r["periodo_vacacional"],
                    "dias_maximos": DIAS_MAXIMOS_POR_PERIODO,
                    "dias_usados": 0,
                    "saldo_disponible": 0,
                    "fuente_archivo": r["fuente_archivo"]
                }

            agrupado[clave]["dias_usados"] += r["dias_usados"]

    for clave in agrupado:
        usados = agrupado[clave]["dias_usados"]
        agrupado[clave]["saldo_disponible"] = DIAS_MAXIMOS_POR_PERIODO - usados

    return list(agrupado.values())

def importar_saldos():
    saldos = agrupar_saldos()
    reemplazar_todos_los_saldos(saldos)
    return {
        "ok": True,
        "mensaje": f"Se importaron {len(saldos)} saldos vacacionales."
    }

def main():
    resultado = importar_saldos()
    print(f"✅ {resultado['mensaje']}")

if __name__ == "__main__":
    main()